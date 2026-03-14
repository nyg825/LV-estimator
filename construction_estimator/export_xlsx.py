"""Generate GMP Hard Cost Budget XLSX matching Acama format."""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from construction_estimator.estimator import UNIT_BASED_CODES


# Division number → (CSI label, display name)
DIVISION_LABELS = {
    1: ("20", "GENERAL REQUIREMENTS"),
    2: ("40", "ON-SITE CONSTRUCTION"),
    3: ("50", "STRUCTURAL CONCRETE"),
    4: ("51", "MASONRY"),
    5: ("52", "STRUCTURAL STEEL & METALS"),
    6: ("53", "ROUGH CARPENTRY & INTERIOR FINISHES"),
    7: ("54", "THERMAL & MOISTURE PROTECTION"),
    8: ("55", "DOORS, WINDOWS & GLAZING"),
    9: ("56", "FINISHES"),
    10: ("57", "SPECIALTIES"),
    11: ("58", "APPLIANCES & EQUIPMENT"),
    12: ("59", "FURNISHINGS"),
    13: ("60", "SPECIAL CONSTRUCTION"),
    14: ("61", "CONVEYING SYSTEMS"),
    15: ("62", "MECHANICAL"),
    16: ("63", "ELECTRICAL"),
    97: ("97", "CONTINGENCY"),
    98: ("98", "GENERAL CONDITIONS"),
    99: ("75", "PROJECT ADMINISTRATION"),
}


def _pricing_note(item, gba, units):
    """Generate a Notes/Source string for a line item."""
    method = getattr(item, "method", "")
    total = item.estimated_total

    if method == "percentage":
        return item.description  # already contains "6%" etc.
    elif method == "hvac_rates":
        if units > 0:
            per_unit = total / units
            return f"{per_unit:,.2f} $/unit \u00d7 {units:,} units \u2014 HVAC rates"
        return "HVAC rates"
    elif method == "elevator_calc":
        return f"Elevator calc \u2014 count \u00d7 stops \u00d7 $31,400"
    elif method == "construction_elevator":
        return "Fixed \u2014 construction elevator"
    elif method == "shoring_calc":
        return f"$96/SF shoring"
    elif method == "structural_concrete":
        return f"$45/SF structural concrete"
    elif method == "per_unit":
        if units > 0:
            per_unit = total / units
            return f"{per_unit:,.2f} $/unit \u00d7 {units:,} units"
        return "per unit"
    elif method == "per_sf":
        if gba > 0:
            per_sf = total / gba
            return f"{per_sf:.3f} $/SF \u00d7 {gba:,.0f} SF"
        return "per SF"
    else:
        if gba > 0:
            per_sf = total / gba
            return f"{per_sf:.3f} $/SF \u00d7 {gba:,.0f} SF"
        return ""


def generate_gmp_xlsx(estimate, form_data):
    """Create an XLSX workbook matching the Acama GMP format.

    Args:
        estimate: Estimate dataclass from estimator
        form_data: dict of form inputs (gba_concrete, gba_wood, unit_mix, etc.)

    Returns:
        BytesIO buffer containing the XLSX file
    """
    wb = Workbook()

    # ── Styles ──
    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    blue_font = Font(name="Arial", size=10, color="0000FF")
    bold_font = Font(name="Arial", size=10, bold=True)
    div_header_font = Font(name="Arial", size=10, bold=True)
    div_header_fill = PatternFill("solid", fgColor="D9E2F3")
    total_fill = PatternFill("solid", fgColor="E2EFDA")
    grand_total_fill = PatternFill("solid", fgColor="C6EFCE")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    currency_fmt = '#,##0'
    currency_fmt2 = '#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style="thin", color="AAAAAA"),
    )
    thick_border = Border(
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    gba = estimate.target_gba
    units = estimate.target_units
    gba_concrete = form_data.get("gba_concrete", 0)
    gba_wood = form_data.get("gba_wood", 0)
    podium_levels = form_data.get("podium_levels", 0)
    wood_levels = form_data.get("wood_levels", 4)
    unit_mix = form_data.get("unit_mix", {})
    elevator_count = form_data.get("elevator_count", 1)
    elevator_stops = form_data.get("elevator_stops", 7)
    lot_size = form_data.get("lot_size", 0)
    shored_area = form_data.get("shored_area", 0)
    project_name = form_data.get("project_name", "New Project")

    # ═══════════════════════════════════════════
    # Sheet 1: Project Data
    # ═══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Project Data"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    r = 1
    ws1.cell(r, 1, f"{project_name} \u2014 PROJECT DATA").font = title_font
    r += 2

    ws1.cell(r, 1, "PROJECT IDENTIFICATION").font = header_font
    r += 1
    data_rows = [
        ("Project Name", project_name),
        ("General Contractor", "LV Construction"),
        ("Estimate Date", date.today().strftime("%B %d, %Y")),
    ]
    for label, val in data_rows:
        ws1.cell(r, 1, label).font = normal_font
        ws1.cell(r, 2, val).font = blue_font
        r += 1

    r += 1
    ws1.cell(r, 1, "BUILDING CONFIGURATION").font = header_font
    r += 1
    total_floors = podium_levels + wood_levels
    has_podium = podium_levels > 0 and gba_concrete > 0
    config_rows = [
        ("Number of Stories", total_floors),
        ("Concrete (Podium) Levels", podium_levels),
        ("Wood Frame Levels", wood_levels),
        ("Construction Type", "Type I/III Mixed" if has_podium else "Type III Wood"),
        ("PTAC", "Yes"),
    ]
    for label, val in config_rows:
        ws1.cell(r, 1, label).font = normal_font
        ws1.cell(r, 2, val).font = blue_font
        r += 1

    r += 1
    ws1.cell(r, 1, "GROSS BUILDING AREA").font = header_font
    r += 1
    if gba_concrete > 0:
        ws1.cell(r, 1, "Concrete / Podium Area").font = normal_font
        c = ws1.cell(r, 2, gba_concrete)
        c.font = blue_font
        c.number_format = '#,##0'
        r += 1
    ws1.cell(r, 1, "Wood Frame Area" if gba_concrete > 0 else "Total Gross SF").font = normal_font
    c = ws1.cell(r, 2, gba_wood if gba_wood > 0 else gba)
    c.font = blue_font
    c.number_format = '#,##0'
    r += 1
    ws1.cell(r, 1, "Total Gross SF").font = bold_font
    c = ws1.cell(r, 2, gba)
    c.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    c.number_format = '#,##0'
    r += 2

    ws1.cell(r, 1, "UNIT MIX").font = header_font
    r += 1
    for br_type, count in sorted(unit_mix.items()):
        if count > 0:
            ws1.cell(r, 1, f"{br_type}").font = normal_font
            ws1.cell(r, 2, count).font = blue_font
            r += 1
    ws1.cell(r, 1, "Total Units").font = bold_font
    c = ws1.cell(r, 2, units)
    c.font = Font(name="Arial", size=10, bold=True, color="0000FF")
    r += 2

    ws1.cell(r, 1, "SYSTEMS").font = header_font
    r += 1
    sys_rows = [
        ("Elevator Count", elevator_count),
        ("Elevator Stops", elevator_stops),
    ]
    if lot_size > 0:
        sys_rows.append(("Lot Size (SF)", lot_size))
    if shored_area > 0:
        sys_rows.append(("Shored Area (SF)", shored_area))
    for label, val in sys_rows:
        ws1.cell(r, 1, label).font = normal_font
        ws1.cell(r, 2, val).font = blue_font
        r += 1

    # ═══════════════════════════════════════════
    # Sheet 2: GMP Hard Cost
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("GMP Hard Cost")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 45
    ws2.column_dimensions["F"].width = 14

    # Title rows
    r = 1
    ws2.cell(r, 1, f"{project_name} \u2014 GMP HARD COST BUDGET").font = title_font
    r += 1
    ws2.cell(r, 1, "LV Construction").font = normal_font
    ws2.cell(r, 4, "Estimate Date:").font = normal_font
    ws2.cell(r, 5, date.today().strftime("%B %d, %Y")).font = blue_font
    r += 1

    # Column headers
    headers = ["Code", "Description", "Allowance", "Amount ($)", "Notes / Source"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws2.cell(r, c_idx, h)
        cell.font = header_font
        cell.alignment = center
        cell.border = Border(bottom=Side(style="medium"))
    r += 1

    # Track trade subtotal row for reference
    trade_subtotal_row = None

    # Helper to write a division header
    def write_div_header(row, label):
        cell = ws2.cell(row, 1, label)
        cell.font = div_header_font
        cell.fill = div_header_fill
        for ci in range(1, 6):
            ws2.cell(row, ci).fill = div_header_fill
        return row + 1

    # Helper to write a line item
    def write_line_item(row, item, is_allowance=False):
        ws2.cell(row, 1, item.cost_code).font = normal_font
        ws2.cell(row, 2, item.description).font = normal_font
        if is_allowance:
            ws2.cell(row, 3, "allowance").font = Font(name="Arial", size=9, italic=True, color="666666")
        c = ws2.cell(row, 4, item.estimated_total)
        c.font = normal_font
        c.number_format = currency_fmt
        c.alignment = right_align
        note = _pricing_note(item, gba, units)
        ws2.cell(row, 5, note).font = Font(name="Arial", size=9, color="555555")
        ws2.cell(row, 5).alignment = wrap
        return row + 1

    # Helper to write division total
    def write_div_total(row, label, first_row, last_row):
        ws2.cell(row, 2, label).font = bold_font
        c = ws2.cell(row, 4)
        c.value = f"=SUM(D{first_row}:D{last_row})"
        c.font = bold_font
        c.number_format = currency_fmt
        c.alignment = right_align
        for ci in range(1, 6):
            ws2.cell(row, ci).fill = total_fill
        return row + 2  # blank row after

    # Compute trade subtotal (all divisions except 97, 98, 99)
    trade_divs = [d for d in estimate.divisions if d.number not in (97, 98, 99)]
    trade_sub = sum(d.estimated_total for d in trade_divs)

    # Trade Subtotal reference row
    ws2.cell(r, 2, "Trade Subtotal (base for % calcs)").font = Font(name="Arial", size=10, bold=True, italic=True)
    c = ws2.cell(r, 6, trade_sub)
    c.font = Font(name="Arial", size=10, bold=True, italic=True)
    c.number_format = currency_fmt
    trade_subtotal_row = r
    r += 1

    # Write each division
    div_total_refs = []  # track (div_num, total_row) for grand total formula

    for div in estimate.divisions:
        div_num = div.number
        label_info = DIVISION_LABELS.get(div_num, (str(div_num), div.name))
        csi_code, display_name = label_info

        r = write_div_header(r, f"{display_name} (CSI {csi_code})" if div_num <= 16 else display_name)

        first_item_row = r
        if div.line_items:
            for item in div.line_items:
                is_allow = "allowance" in (item.description or "").lower() or "fixed" in (getattr(item, "method", "") or "").lower()
                r = write_line_item(r, item, is_allowance=is_allow)
        else:
            # Division with no line items (percentage-based like contingency/GC)
            # Write as a single row
            from construction_estimator.models import EstimateLineItem
            pseudo = EstimateLineItem(
                cost_code="",
                description=div.name,
                division_number=div_num,
                division_name=div.name,
                estimated_total=div.estimated_total,
                estimated_per_sf=div.estimated_per_sf,
                estimated_per_unit=div.estimated_per_unit,
                low_total=div.low_total,
                high_total=div.high_total,
                confidence=1.0,
                data_points=0,
                method="percentage",
            )
            r = write_line_item(r, pseudo)

        last_item_row = r - 1
        div_total_label = f"{display_name.split('(')[0].strip()} Total"
        r = write_div_total(r, div_total_label, first_item_row, last_item_row)
        div_total_refs.append((div_num, r - 2))  # row of the total

    # Grand total
    ws2.cell(r, 2, "TOTAL HARD COSTS \u2014 GMP").font = Font(name="Arial", size=12, bold=True)
    # Sum all division totals
    total_refs = "+".join(f"D{tr}" for _, tr in div_total_refs)
    c = ws2.cell(r, 4)
    c.value = estimate.project_total
    c.font = Font(name="Arial", size=12, bold=True)
    c.number_format = currency_fmt
    c.alignment = right_align
    for ci in range(1, 6):
        ws2.cell(r, ci).fill = grand_total_fill
        ws2.cell(r, ci).border = thick_border
    r += 2

    # Summary metrics
    ws2.cell(r, 1, "Gross SF").font = bold_font
    c = ws2.cell(r, 2, gba)
    c.font = blue_font
    c.number_format = '#,##0'
    r += 1
    ws2.cell(r, 1, "Total Units").font = bold_font
    c = ws2.cell(r, 2, units)
    c.font = blue_font
    r += 1
    ws2.cell(r, 1, "Total Hard Cost").font = bold_font
    c = ws2.cell(r, 2, estimate.project_total)
    c.font = bold_font
    c.number_format = currency_fmt
    r += 1
    ws2.cell(r, 1, "Cost per SF ($/SF)").font = bold_font
    c = ws2.cell(r, 2, estimate.cost_per_sf)
    c.font = bold_font
    c.number_format = currency_fmt2
    r += 1
    ws2.cell(r, 1, "Cost per Unit ($/Unit)").font = bold_font
    c = ws2.cell(r, 2, estimate.cost_per_unit)
    c.font = bold_font
    c.number_format = currency_fmt2
    r += 2

    # Legend
    ws2.cell(r, 1, "LEGEND:").font = Font(name="Arial", size=9, bold=True)
    ws2.cell(r, 2, "Blue text = hardcoded input  |  Black text = formula  |  Allowance = confirmed scope item with estimated cost").font = Font(name="Arial", size=9, color="666666")

    # Print settings
    ws2.sheet_properties.pageSetUpPr = None
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
